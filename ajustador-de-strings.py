# Importação de pacotes
import pandas as pd
from rapidfuzz import process, fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import re
import os
import unicodedata

# Configurações gerais- A serem alteradas pelos usuários

#Ifood
#caminho_planilha = 'C:/Users/OPB/Desktop/Python/Ifood/bd_ifood_VFINAL.xlsx'
#nome_aba = 'bd'
#colunas_analise = ['Q3_1', 'Q3_2','Q3_3','Q3_4','Q3_5']  
#percentual_similaridade_minimo = 80
#aba_alvo = 'Delivery_Bebidas'

#Zé delivery
#caminho_planilha = 'C:/Users/OPB/Desktop/Python/Ze_delivery/bd_ze_delivery_Jun25.xlsx'
#nome_aba = 'bd'
#colunas_analise = ['Q8', 'Q9']  
#percentual_similaridade_minimo = 80
#aba_alvo = 'Delivery_Bebidas'

#CRM
#caminho_planilha = 'C:/Users/OPB/Desktop/Python/Grupo CRM/bd_consolidado 18jul.xlsx'
#nome_aba = 'bd'
#colunas_analise = ['Q5', 'Q6_1_TEXT','Q6_2_TEXT', 'Q6_3_TEXT','Q6_4_TEXT','Q6_5_TEXT','Q58']  
#percentual_similaridade_minimo = 80
#aba_alvo = 'Loja_chocolate'

#Mobile Time- Uso de apps
#caminho_planilha = 'C:/Users/OPB/Desktop/Python/Mobile Time/bd_consistencia_Uso de apps.xlsx'
#ome_aba = 'bd'
#colunas_analise = ['Q7_1', 'Q7_2', 'Q7_3','Q7_4','Q7_5', 'Q7_6', 'Q7_7',
#                   'Q7_8','Q7_9','Q7_10','Q7_11','Q7_12','Q7_13','Q7_14',
#                  'Q7_15','Q7_16','Q7_17','Q7_18','Q7_19','Q7_20','Q7_1',
#                   'Q8','Q9', 'Q10','Q12','Q25','Q27','Q32','Q34','Q36','Q37',
#                  'Q39','Q41','Q44','Q55']  

#percentual_similaridade_minimo = 80
#aba_alvo = 'Apps' 

caminho_planilha = input("Insira o caminho da planilha a ser processada: ")
nome_aba = input("Insira a aba da planilha a ser processada: ")
entrada_colunas_analise = input("Insira as colunas a serem processadas separadas por ponto e vírgula: EX: Q1;Q2 ")
colunas_analise = entrada_colunas_analise.replace(' ', '').split(';')
percentual_similaridade_minimo = input("Insira o percentual mínimo de similaridade entre as palavras que serão ajustadas: APENAS INTEIROS, EX: 80. : ")
caminho_alvos = input("Insira o caminho da planilha de alvos: ")
aba_alvo = input("Insira o nome da aba da planilha de alvos correspondente ao projeto: ")

# Leitura da lista de alvos a partir da planilha
df_alvos = pd.read_excel(caminho_alvos, sheet_name=aba_alvo)
alvos = df_alvos['alvo'].dropna().tolist()  # Remove valores vazios

# Carrega o arquivo Excel
data_frame = pd.read_excel(caminho_planilha, sheet_name=nome_aba, engine='openpyxl')
bd = data_frame.copy()  #copia de data_frame

# Gera caminho de saída automaticamente adicionando "_corrigido" antes da extensão
base, ext = os.path.splitext(caminho_planilha)
caminho_saida = f"{base}_corrigido{ext}"

# Estilos de destaque
fill = PatternFill(start_color='FFFFCCCC', end_color='FFFFCCCC', fill_type='solid')
font = Font(color='9C0006')

# Dicionário para armazenar os resultados
resultados_correcao = {}
destaques = {}

# Lista de expressões a padronizar como "Não sei/Não lembro"
nao_sei_variacoes = [
    'não lembro', 'nao lembro', 'não sei', 'nao sei', 'não uso', 'nao uso',
    'não', 'nao', 'nenhum', 'nenhuma', 'nemhuma', 'nem uma', 'nao conheco',
    'não tenho ideia', 'não peço', 'não utilizo', 'nao utilizo', 'nao conheço',
    'nao sei nenhum', 'nao tenho ideia' 'todas', 'todos', 'nao me lembro',
    'não gravei na memori','não gravei na memoria','nao gravei na memoria',
    'várias','varias','que?','nem um','sei la','sei lá','nenhu', 'varios jogos', 
    'varios','diversos','diversas','outros jogos','outros aplicativos','todo','todos',
    'toda','todas','n sei','ns','varios jagos ofline','Ñ sei 5','Ñ sei','app','apps',
    'apss','esqueci','esqueci o nome','ferramenta?','n lembro','na lembro','jogo', 'não tenho ideia',
    'Não tenho app', 'Não peço', 'Nem uma', 'Nao conheco', 'Não utilizo', 'E difícil eu usar aplicativo'
]

# Função para normalizar texto
def normalizar(texto):
    if pd.isna(texto):
        return ''
    texto = str(texto).lower().strip()
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode()
    texto = re.sub(r'[^a-z0-9]', '', texto)  # mantém letras e números, remove o resto
    return texto

# Loop principal
for coluna in colunas_analise:
    correcao_final = []
    destacar_vermelho = []

    for idx, valor in enumerate(bd[coluna]):
        if idx < 2:
            correcao_final.append(valor)
            destacar_vermelho.append(False)
            continue
        # Garante que 'valor' é string, ou string vazia se for NaN
        original = str(valor) if pd.notna(valor) else ""

        # Separa por vírgulas, barras, hífens, etc. – só a primeira parte
        original_base = re.split(r'[.,/;_]', original)[0].strip()

        normalizado = normalizar(original_base)
        corrigido = original_base
        marcado = True
        marcado = bool(original.strip())  

    # Etapa 1: Verifica se é o caso especial de "Zé delivery"
        if aba_alvo == 'Delivery_Bebidas' and re.search(r'\bz[é|e]\b', normalizado):
            corrigido = 'Zé delivery'
            marcado = False

        elif aba_alvo == 'Delivery_Bebidas' and re.search(r'\bbk\b', normalizado):
            corrigido = 'Burger King'
            marcado = False

        elif aba_alvo == 'Loja_chocolate' and re.search(r'\b(cacau.*brasil|brasil.*cacau)\b', original, re.IGNORECASE):
            corrigido = 'Brasil Cacau'
            marcado = False

        elif aba_alvo == 'Loja_chocolate' and re.search(r'\b(show)\b', normalizado):
            corrigido = 'Cacau Show'
            marcado = False
        
        else:
            # Etapa 2: Fuzz com alvos
            alvos_normalizados = [normalizar(a) for a in alvos]
            mapa_alvo_original = dict(zip(alvos_normalizados, alvos))

            resultado = process.extractOne(
                normalizado,
                alvos_normalizados,
                scorer = fuzz.WRatio,
                score_cutoff = int(percentual_similaridade_minimo)
            )

            if resultado is not None:
                melhor_alvo_normalizado, _ , _ = resultado
                corrigido = mapa_alvo_original[melhor_alvo_normalizado]
                marcado = False

            # Etapa 3: Verifica se o valor está na lista de variações
            if len(normalizado) == 1 or normalizado in nao_sei_variacoes:
                corrigido = 'Não sei/Não lembro'
                marcado = False
                
         # Regras específicas para Apps (aplicadas após o Fuzz)
            if aba_alvo in ['Apps', 'mobile_time']:
                padroes_apps = [
                    (r'\brenner\b', 'Lojas Renner'),
                    (r'\bmercantil\b', 'Banco Mercantil'),
                    (r'\binter\b', 'Banco Inter'),
                    (r'\b(x\s*\/\s*twitter|x\s+twitter|twitter\s*\/\s*x|x|twitter)\b', 'X/Twitter'),
                    (r'\bgov\b', 'gov.br'),
                    (r'\bchrome\b|\bnavegador\s+(google|chrome)\b', 'Google Chrome'),
                    (r'\b(watts?|wats?|whats?|wpp?|wattzap?|zap+|zapp?|watzap?|whts?|atzap?|wattzapp?|wattszap?|wstsap?|watzp?|wastszap?|wha?|wwastzap?)\b', 'Whatsapp'),
                    (r'\bsms\b', 'Mensagem'),
                    (r'\bb[ií]blia(\s+sagrada)?\b', 'Bíblia'),
                    (r'\bxp\b', 'XP Investimentos'),
                    (r'\bbrb\b', 'Banco BRB'),
                    (r'\b(face|feice|fece)\b', 'Facebook'),
                    (r'\b(insta)\b', 'Instagram'),
                    (r'\bmequi\b', "McDonald's"),
                    (r'\blol\b', 'League of Legends')
                ]
            
                for padrao, nome_corrigido in padroes_apps:
                    if re.search(padrao, normalizado):
                        corrigido = nome_corrigido
                        marcado = False
                        break  # para evitar sobrescrever caso mais de um padrão bata

        correcao_final.append(corrigido)
        destacar_vermelho.append(marcado)

    nova_coluna = f'{coluna}_corrigida'
    bd[nova_coluna] = correcao_final
    resultados_correcao[coluna] = nova_coluna
    destaques[coluna] = destacar_vermelho

# Salva o DataFrame no Excel
bd.to_excel(caminho_saida, index=False)

# Aplica destaque com openpyxl
wb = load_workbook(caminho_saida)
ws = wb.active

for coluna in colunas_analise:
    nova_coluna = resultados_correcao[coluna]
    col_idx = list(bd.columns).index(nova_coluna) + 1  

    for i, marcar in enumerate(destaques[coluna], start=1):  
        linha_excel = i + 1 
        if i >= 2 and marcar:
            cell = ws.cell(row=linha_excel, column=col_idx)
            cell.fill = fill
            cell.font = font

wb.save(caminho_saida)
wb.close()